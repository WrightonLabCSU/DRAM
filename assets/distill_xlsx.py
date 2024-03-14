import pandas as pd
import sqlite3
import argparse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
import re

import logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def parse_arguments():
    parser = argparse.ArgumentParser(description='Generate a multi-sheet XLSX document from distill sheets and a SQLite database.')
    parser.add_argument('--target_id_counts', type=str, help='Path to the target_id_counts.tsv file.')
    parser.add_argument('--db_name', type=str, help='Name of the SQLite database file.')
    parser.add_argument('--distill_sheets', nargs='+', help='List of paths to distill sheets.')
    parser.add_argument('--rrna_file', type=str, default=None, help='Path to the rrna_sheet.tsv file.')
    parser.add_argument('--combined_rrna_file', type=str, default=None, help='Path to the combined rRNA TSV file.')
    parser.add_argument('--trna_file', type=str, default=None, help='Path to the trna_sheet.tsv file.')
    parser.add_argument('--output_file', type=str, help='Path to the output XLSX file.')
    parser.add_argument('--quast', type=str, default=None, help='Path to the QUAST stats TSV file.')
    return parser.parse_args()

def sum_counts_for_multi_gene_ids(target_id_counts_df, gene_ids):
    """Sum counts for each gene ID across all samples."""
    summed_counts = {col: 0 for col in target_id_counts_df.columns if col != 'gene_id'}
    for gene_id in gene_ids:
        counts = target_id_counts_df[target_id_counts_df['gene_id'] == gene_id]
        for col in summed_counts:
            summed_counts[col] += counts[col].sum() if not counts.empty else 0
    return summed_counts

def sum_counts_for_gene_id(gene_id, target_id_counts_df, aggregated_counts):
    if gene_id in target_id_counts_df['gene_id'].values:
        gene_counts = target_id_counts_df[target_id_counts_df['gene_id'] == gene_id]
        for col in target_id_counts_df.columns[1:]:  # Skip the gene_id column
            if col not in aggregated_counts:
                aggregated_counts[col] = 0
            aggregated_counts[col] += gene_counts[col].sum()

def compile_target_id_counts(target_id_counts):
    """Compile target ID counts from the specified TSV file."""
    return pd.read_csv(target_id_counts, sep='\t')

def compile_quast_info(quast_file):
    """Compile QUAST info if the file contains data."""
    if not file_contains_data(quast_file):
        print(f"Skipping QUAST processing for {quast_file} as it contains 'NULL'.")
        return None
    return pd.read_csv(quast_file, sep='\t')

def read_distill_sheets(distill_sheets):
    """Read distill sheets and compile them into a dictionary."""
    sheets_data = {}
    for sheet_path in distill_sheets:
        if file_contains_data(sheet_path):
            df = pd.read_csv(sheet_path, sep='\t')
            topic = df['topic_ecosystem'].unique().tolist()
            column_type = 'ec_id' if 'ec_id' in df.columns else 'gene_id'
            sheets_data.update({sheet_path: {'dataframe': df, 'topics': topic, 'column_type': column_type}})
        else:
            print(f"Skipping {sheet_path} as it contains 'NULL'.")
    return sheets_data

def file_contains_data(file_path):
    """Check if the file contains data other than 'NULL'."""
    try:
        with open(file_path, 'r') as f:
            first_line = f.readline().strip()
            return first_line != "NULL"
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return False

def fetch_all_gene_ids(db_name):
    """
    Fetch all unique gene IDs from the annotations database.
    """
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT gene_id FROM annotations")
    all_gene_ids = [row[0] for row in cursor.fetchall()]
    conn.close()
    return all_gene_ids

def filter_and_aggregate_counts(gene_ids, target_id_counts_df, db_name, all_gene_ids_in_db):
    """
    Filters gene_ids to those found in the annotations database and aggregates their counts.
    """
    filtered_gene_ids = []
    aggregated_counts = {col: 0 for col in target_id_counts_df.columns if col != 'gene_id'}

    for gene_id in gene_ids:
        if gene_id in all_gene_ids_in_db or is_partial_ec_number(gene_id):
            filtered_gene_ids.append(gene_id)
            if is_partial_ec_number(gene_id):
                matching_ec_numbers = fetch_matching_ec_numbers(db_name, gene_id)
                for ec_number in matching_ec_numbers:
                    sum_counts_for_gene_id(ec_number, target_id_counts_df, aggregated_counts)
            else:
                sum_counts_for_gene_id(gene_id, target_id_counts_df, aggregated_counts)

    return filtered_gene_ids, aggregated_counts


def fetch_matching_ec_numbers(db_name, partial_ec_number):
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    
    # Use the SQL pattern for matching any EC numbers that start with the partial EC number
    partial_ec_pattern = partial_ec_number.replace("EC:", "").rstrip("-") + "%"
    
    cursor.execute("SELECT DISTINCT gene_id FROM annotations WHERE gene_id LIKE ?", ("%" + partial_ec_pattern,))
    all_rows = cursor.fetchall()
    
    
    matching_ec_numbers = []
    for row in all_rows:
        # Split each row by common separators and remove any leading/trailing whitespace
        for gene_id in re.split('; |, |;|,', row[0]):
            gene_id = gene_id.strip()
            # Add gene_id to the set if it matches the partial EC pattern
            if gene_id.startswith("EC:" + partial_ec_pattern.rstrip("%")):
                matching_ec_numbers.append(gene_id)
                
    conn.close()
    return matching_ec_numbers

def aggregate_counts(gene_ids, target_id_counts_df, db_name):
    aggregated_counts = {col: 0 for col in target_id_counts_df.columns if col != 'gene_id'}

    if is_partial_ec_number(gene_ids[0]):  # Assuming all gene_ids are either all partial or all direct EC numbers
        partial_matches = fetch_matching_ec_numbers(db_name, gene_ids[0])

        # Generate regex pattern to match any of the partial EC numbers within a gene_id string
        partial_matches_pattern = '|'.join(map(re.escape, partial_matches))

        # Filter target_id_counts_df for rows where gene_id matches any of the partial EC numbers
        for _, row in target_id_counts_df.iterrows():
            if re.search(partial_matches_pattern, row['gene_id']):
                for col in aggregated_counts.keys():
                    aggregated_counts[col] += row[col]

    else:
        # Handle direct gene ID matches
        for gene_id in gene_ids:
            if gene_id in target_id_counts_df['gene_id'].values:
                match_counts = target_id_counts_df.loc[target_id_counts_df['gene_id'] == gene_id]
                for col in aggregated_counts.keys():
                    aggregated_counts[col] += match_counts[col].sum()

    return aggregated_counts

def process_distill_sheet_topic(df_topic, target_id_counts_df, db_name):
    """
    Process each topic within a distill sheet for composite gene_id entries and partial EC numbers.
    """
    processed_rows = []
    any_gene_identified = False

    for _, row in df_topic.iterrows():
        gene_ids = split_gene_ids(row['gene_id'])
        valid_gene_ids = [gene_id for gene_id in gene_ids if gene_id in fetch_all_gene_ids(db_name) or is_partial_ec_number(gene_id)]

        if not valid_gene_ids:
            continue

        # Check if any gene IDs are partial EC numbers
        partial_ec_gene_ids = [gene_id for gene_id in valid_gene_ids if is_partial_ec_number(gene_id)]
        regular_gene_ids = [gene_id for gene_id in valid_gene_ids if not is_partial_ec_number(gene_id)]

        # Call filter_and_aggregate_counts for regular gene IDs
        if regular_gene_ids:
            _, aggregated_counts_regular = filter_and_aggregate_counts(regular_gene_ids, target_id_counts_df, db_name, fetch_all_gene_ids(db_name))
            any_gene_identified = True
        else:
            aggregated_counts_regular = {}

        # Call aggregate_counts for partial EC numbers
        if partial_ec_gene_ids:
            aggregated_counts_partial_ec = aggregate_counts(partial_ec_gene_ids, target_id_counts_df, db_name)
            any_gene_identified = True
        else:
            aggregated_counts_partial_ec = {}

        # Combine aggregated counts
        aggregated_counts = {**aggregated_counts_regular, **aggregated_counts_partial_ec}

        updated_row = row.to_dict()
        for sample_col, count in aggregated_counts.items():
            updated_row[sample_col] = count
        processed_rows.append(updated_row)

    if not any_gene_identified:
        placeholder_row = {"gene_id": "No genes identified for this distill sheet"}
        for col in target_id_counts_df.columns[1:]:
            placeholder_row[col] = 0
        processed_rows = [placeholder_row]

    return pd.DataFrame(processed_rows)


def compile_genome_stats(db_name):
    conn = sqlite3.connect(db_name)
    cursor = conn.cursor()
    
    # Check for the presence of optional columns in the database
    cursor.execute("PRAGMA table_info(annotations)")
    columns_info = cursor.fetchall()
    column_names = [info[1] for info in columns_info]
    
    # Start with a query to select distinct samples
    query_base = "SELECT DISTINCT sample FROM annotations"
    df_samples = pd.read_sql_query(query_base, conn)
    
    # Initialize the genome_stats DataFrame with sample names
    df_genome_stats = pd.DataFrame({"sample": df_samples['sample']})
    
    # Dynamically build a query to include optional columns if they are present
    optional_columns = ['taxonomy', 'Completeness', 'Contamination']
    select_clauses = []
    for col in optional_columns:
        if col in column_names:
            # Prepare a SELECT clause to calculate average for numeric columns and group_concat for taxonomy
            if col in ['Completeness', 'Contamination']:
                select_clauses.append(f"AVG({col}) AS {col}")
            else:  # Assuming 'taxonomy' is not a numeric column
                select_clauses.append(f"GROUP_CONCAT(DISTINCT {col}) AS {col}")
    
    # If there are optional columns to include, modify the base query
    if select_clauses:
        query = f"SELECT sample, {', '.join(select_clauses)} FROM annotations GROUP BY sample"
        df_stats = pd.read_sql_query(query, conn)
        df_genome_stats = pd.merge(df_genome_stats, df_stats, on="sample", how="left")
    else:
        # If no optional columns, the genome_stats will only contain sample names at this point
        pass
    
    conn.close()
    return df_genome_stats

def add_sheet_from_dataframe(wb, df, sheet_name):
    ws = wb.create_sheet(title=sheet_name)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

def prepare_ec_like_patterns(ec_number):
    """
    Prepare SQL LIKE patterns for partial matching of EC numbers.
    Handles cases where EC numbers are partial or contain multiple EC numbers separated by semicolons or spaces.
    """
    patterns = []
    # Split multiple EC numbers and prepare patterns for each
    parts = ec_number.replace(" ", ";").split(";")  # Split by semicolon and space
    for part in parts:
        clean_part = part.strip().rstrip('.')
        if clean_part:  # Ensure the part is not empty after stripping
            # Append '%' to match any characters following the specified EC number part
            pattern = clean_part + "%"
            patterns.append(pattern)
    return patterns

def query_annotations_for_gene_ids(db_name, ids):
    conn = sqlite3.connect(db_name)
    df_result = pd.DataFrame(columns=['gene_id'])
    all_gene_ids = pd.read_sql_query("SELECT DISTINCT gene_id FROM annotations", conn)

    for id_value in ids:
        # Split based on various separators, and strip whitespace
        split_ids = re.split('; |, |;|,', id_value)
        split_ids = [id.strip() for id in split_ids if id.strip()]
        
        matches_found = []  # To keep track of found matches for composite gene_ids
        for part_id in split_ids:
            # Search for each part in the database
            if all_gene_ids['gene_id'].str.contains(f'^{part_id}$').any():
                matches_found.append(part_id)

        if matches_found:
            # If any part of the composite gene_id is found, include the original composite gene_id in the results
            df_result = pd.concat([df_result, pd.DataFrame({'gene_id': [id_value]})], ignore_index=True)

    df_result.drop_duplicates(inplace=True)
    conn.close()
    return df_result

def compile_rrna_information(combined_rrna_file):
    """Compile rRNA information from the combined rRNA file."""
    rrna_data = pd.read_csv(combined_rrna_file, sep='\t')
    # Group by sample and type to concatenate query_id and positions
    rrna_summary = rrna_data.groupby(['sample', 'type']).apply(
        lambda x: '; '.join([f"{row['query_id']} ({row['begin']}, {row['end']})" for _, row in x.iterrows()])
    ).unstack(fill_value='')
    rrna_summary.reset_index(inplace=True)
    return rrna_summary

def add_rrna_trna_sheets(wb, rrna_file, trna_file):
    for tsv_file, sheet_name in [(rrna_file, 'rRNA'), (trna_file, 'tRNA')]:
        if tsv_file and file_contains_data(tsv_file):
            df = pd.read_csv(tsv_file, sep='\t')
            add_sheet_from_dataframe(wb, df, sheet_name)

def compile_rrna_data(rrna_file):
    """Compile rRNA data from the given file."""
    rrna_data = pd.read_csv(rrna_file, sep='\t')
    # Process rRNA data as required for your use case
    # This might involve summarizing the rRNA types and locations for each sample
    return rrna_data

def compile_trna_counts(trna_file):
    trna_data = pd.read_csv(trna_file, sep='\t')
    sample_columns = trna_data.columns[5:]  # Adjust based on actual structure
    trna_counts = pd.DataFrame([{'sample': sample, 'tRNA count': trna_data[sample].sum()} for sample in sample_columns])
    return trna_counts

def update_genome_stats_with_rrna_trna(genome_stats_df, rrna_file, trna_file):
    """Update the genome_stats DataFrame with rRNA and tRNA information."""
    # Process rRNA file if it contains data
    if file_contains_data(rrna_file):
        rrna_data = compile_rrna_data(rrna_file)
        # Integrate rrna_data into genome_stats_df as needed

    # Process tRNA file if it contains data
    if file_contains_data(trna_file):
        trna_counts = compile_trna_counts(trna_file)
        # Merge tRNA counts into genome_stats_df
        genome_stats_df = pd.merge(genome_stats_df, trna_counts, on="sample", how="left")

    return genome_stats_df

def update_genome_stats_with_rrna(genome_stats_df, combined_rrna_file):
    """Update the genome_stats DataFrame with rRNA information if available."""
    if file_contains_data(combined_rrna_file):
        rrna_summary = compile_rrna_information(combined_rrna_file)
        genome_stats_df = pd.merge(genome_stats_df, rrna_summary, on="sample", how="left")
    return genome_stats_df

def add_optional_sheets(wb, args):
    """Adds optional rRNA and tRNA sheets if data is available."""
    if args.rrna_file and file_contains_data(args.rrna_file):
        add_sheet_from_dataframe(wb, pd.read_csv(args.rrna_file, sep='\t'), 'rRNA')
    
    if args.trna_file and file_contains_data(args.trna_file):
        add_sheet_from_dataframe(wb, pd.read_csv(args.trna_file, sep='\t'), 'tRNA')

def is_partial_ec_number(gene_id):
    is_partial = gene_id.startswith("EC:") and gene_id.endswith("-")
    return is_partial

def split_gene_ids(gene_id_field):
    # This function will split the gene_id field into individual gene IDs or EC numbers
    # Handle different separators and strip whitespace
    return [gene_id.strip() for gene_id in re.split(r'[;,]\s*', gene_id_field)]

def main():
    args = parse_arguments()

    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # Compile genome stats and add as a sheet
    genome_stats_df = compile_genome_stats(args.db_name)
    genome_stats_df = update_genome_stats_with_rrna_trna(genome_stats_df, args.rrna_file, args.trna_file)
    genome_stats_df = update_genome_stats_with_rrna(genome_stats_df, args.combined_rrna_file)
    if args.quast:
        quast_data = compile_quast_info(args.quast)
        if quast_data is not None:
            genome_stats_df = pd.merge(genome_stats_df, quast_data, on="sample", how="left")
    add_sheet_from_dataframe(wb, genome_stats_df, "Genome_Stats")

    # Read target ID counts
    target_id_counts_df = compile_target_id_counts(args.target_id_counts)
    
    # Fetch all gene IDs from the annotations database
    all_gene_ids_in_db = fetch_all_gene_ids(args.db_name)
    
    # Process each distill sheet
    distill_data = read_distill_sheets(args.distill_sheets)

    for sheet_path, info in distill_data.items():
        df_distill = info['dataframe']
        
        for topic in info['topics']:
            df_topic = df_distill[df_distill['topic_ecosystem'] == topic]
            df_topic_final = process_distill_sheet_topic(df_topic, target_id_counts_df, args.db_name)
            
            sheet_name = topic[:31]  # Limit sheet name to 31 characters

            if df_topic_final is None or df_topic_final.empty:
                # No genes identified for this topic
                ws = wb.create_sheet(title=sheet_name)
                ws['A1'] = "No genes identified for this distill sheet"
            else:
                # Add the processed dataframe to the workbook
                add_sheet_from_dataframe(wb, df_topic_final, sheet_name)

    # Add rRNA and tRNA sheets if available
    if args.rrna_file and file_contains_data(args.rrna_file):
        rrna_df = pd.read_csv(args.rrna_file, sep='\t')
        add_sheet_from_dataframe(wb, rrna_df, 'rRNA')

    if args.trna_file and file_contains_data(args.trna_file):
        trna_df = pd.read_csv(args.trna_file, sep='\t')
        add_sheet_from_dataframe(wb, trna_df, 'tRNA')

    # Save the workbook
    wb.save(args.output_file)

if __name__ == '__main__':
    main()
