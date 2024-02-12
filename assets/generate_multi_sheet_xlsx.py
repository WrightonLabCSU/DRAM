import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def is_null_content(file_path):
    with open(file_path, 'r') as file:
        content = file.read().strip()
    return content == "NULL"

def generate_multi_sheet_xlsx(input_file, rrna_file, trna_file, combined_annotations, combined_rrna, output_file):
    # Read the data from the input file using pandas with tab as the separator
    data = pd.read_csv(input_file, sep='\t')

    hardcoded_columns = ["gene_id", "gene_description", "pathway", "topic_ecosystem", "category", "subcategory"]

    # Read combined_annotations
    combined_data = pd.read_csv(combined_annotations, sep='\t')

    # Deduplicate and extract unique sample values
    unique_samples = combined_data['sample'].unique()

    # Create a Workbook
    wb = Workbook()

    # Create the genome_stats sheet
    gs_sheet = wb.create_sheet(title="genome_stats")

    # Dynamically get unique RNA types from combined_rrna
    rrna_data = None  # Initialize rrna_data to None
    unique_rna_types = []

    if not is_null_content(combined_rrna):
        # Dynamically get unique RNA types from combined_rrna
        rrna_data = pd.read_csv(combined_rrna, sep='\t')
        unique_rna_types = rrna_data['type'].unique()

    # Append column names to genome_stats sheet
    column_names = ["sample", "number of scaffolds"]

    # Check if the columns exist in combined_annotations_df and append them if they do
    if "taxonomy" in combined_data.columns:
        column_names.append("taxonomy")
    if "Completeness" in combined_data.columns:
        column_names.append("completeness")
    if "Contamination" in combined_data.columns:
        column_names.append("contamination")

    if not is_null_content(trna_file):
        column_names += list(unique_rna_types) + ["tRNA count"]
    else:
        column_names += list(unique_rna_types)
    gs_sheet.append(column_names)


    # Populate genome_stats sheet with data from combined_annotations
    for sample in unique_samples:
        # Extract information for the current sample from combined_annotations
        sample_info = combined_data[combined_data['sample'] == sample]

        if not sample_info.empty:
            # Get the first row of sample_info (assuming one row per sample)
            sample_info = sample_info.iloc[0]

            # Extract completeness and contamination values
            completeness = sample_info.get('Completeness', None)
            contamination = sample_info.get('Contamination', None)
        else:
            # Handle the case where sample_info is empty (no data for the sample)
            completeness = None
            contamination = None

        # Append data to genome_stats sheet
        gs_data = [sample, None]

        # Check if the columns exist in combined_annotations_df and append them if they do
        if "taxonomy" in combined_data.columns:
            gs_data.append(sample_info.get('taxonomy', None))
        if "Completeness" in combined_data.columns:
            gs_data.append(completeness)
        if "Contamination" in combined_data.columns:
            gs_data.append(contamination)

        gs_data += [None] * (len(unique_rna_types) + 1)
        gs_sheet.append(gs_data)


    # Update RNA columns dynamically
    if rrna_data is not None:
        for rna_type in unique_rna_types:
            # Find the corresponding column index in the genome_stats sheet
            col_idx = column_names.index(rna_type) + 1
            print(f"\nUpdating RNA column: {rna_type}")
            print(f"Column Index in genome_stats: {col_idx}")

            # Iterate over samples
            for idx, sample in enumerate(unique_samples, start=2):
                # Extract relevant data for the current sample and rna_type
                sample_rrna_data = rrna_data[(rrna_data['sample'] == sample) & (rrna_data['type'] == rna_type)]

                # Check if sample_rrna_data is not empty
                if not sample_rrna_data.empty:
                    # Concatenate the values and format them as needed
                    values = [
                        f"{row['query_id']} ({row['begin']}, {row['end']})"
                        for _, row in sample_rrna_data.iterrows()
                    ]

                    # Join multiple values with "; "
                    joined_values = "; ".join(values)

                    # Update the value in the correct cell
                    gs_sheet.cell(row=idx, column=col_idx).value = joined_values
                    print(f"Updated value at row {idx}, column {col_idx}")
                else:
                    print(f"Sample {sample} has no data for {rna_type}")

    # Sum tRNA counts and update the 'tRNA count' column
    if not is_null_content(trna_file):
        trna_data = pd.read_csv(trna_file, sep='\t')

        for idx, sample in enumerate(unique_samples, start=2):
            # Extract relevant data for the current sample
            sample_trna_count = trna_data[sample].sum(numeric_only=True)

            # Update the 'tRNA count' column in the correct cell
            gs_sheet.cell(row=idx, column=len(column_names)).value = sample_trna_count
            print(f"Updated tRNA count value at row {idx}")

    # Create a dictionary to store data for each sheet
    sheet_data = {}

    for _, row in data.iterrows():
        # Split the "sheet" values by "; " and iterate over them
        for sheet_name in row['topic_ecosystem'].split('; '):
            sheet_name = sheet_name.replace(" ", "_")
            if sheet_name not in sheet_data:
                sheet_data[sheet_name] = {
                    'columns': [],  # Store column names for this sheet
                    'data': []  # Store data rows for this sheet
                }

            # Exclude the expected columns and move other columns
            row_data = [row[col] for col in row.index if col not in column_names]

            # Append the modified row to the corresponding sheet's data
            sheet_data[sheet_name]['data'].append(row_data)

            # Collect column names that are not in column_names
            new_columns = [col for col in row.index if col not in column_names]
            sheet_data[sheet_name]['columns'].extend(new_columns)

    # Inside the loop that creates sheets for topic_ecosystem values
    for sheet_name, sheet_info in sheet_data.items():
        # Create a worksheet for each sheet
        ws = wb.create_sheet(title=sheet_name)

        # Extract unique column names for this sheet
        unique_column_names = []  # Start with an empty list to maintain order
        for col in sheet_info['columns']:
            if col not in unique_column_names and col not in hardcoded_columns:
                unique_column_names.append(col)

        # Define the desired order of columns (including hardcoded columns)
        sorted_column_names = hardcoded_columns + unique_column_names

        # Append column names as the first row
        ws.append(sorted_column_names)

        # Print the final column names of this sheet for debugging
        print(f'"{sheet_name}" sheet: Final column names: {sorted_column_names}')

        # Append data rows to the worksheet
        for row in sheet_info['data']:
            # Create a dict from column names to row data for sorting
            row_dict = dict(zip(sheet_info['columns'], row))

            # Sort the row data according to the sorted_column_names
            sorted_row = [row_dict.get(col) for col in sorted_column_names]
            ws.append(sorted_row)

        # Create a table from the data for filtering
        tab = Table(displayName=f"{sheet_name}_Table", ref=ws.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium9", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=True
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

    # Before adding rRNA sheets
    print("Adding rRNA sheet")
    if not is_null_content(rrna_file):
        # Add rRNA sheet
        rrna_data = pd.read_csv(rrna_file, sep='\t')
        rrna_sheet = wb.create_sheet(title="rRNA")

        # Append column names as the first row
        rrna_sheet.append(list(rrna_data.columns))

        # Append data rows to the worksheet
        for _, row in rrna_data.iterrows():
            rrna_sheet.append(list(row))


    print("Adding tRNA sheet")
    if not is_null_content(trna_file):
        # Add tRNA sheet
        trna_data = pd.read_csv(trna_file, sep='\t')
        trna_sheet = wb.create_sheet(title="tRNA")

        # Append column names as the first row
        trna_sheet.append(list(trna_data.columns))

        # Append data rows to the worksheet
        for _, row in trna_data.iterrows():
            trna_sheet.append(list(row))

    # Before removing the default "Sheet" that was created
    print("Removing default 'Sheet'")
    # Remove the default "Sheet" that was created
    default_sheet = wb['Sheet']
    wb.remove(default_sheet)

    # Save the workbook as the output file
    print(f"Saving workbook to {output_file}")
    wb.save(output_file)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Generate multi-sheet XLSX file from TSV files')
    parser.add_argument('--input_file', help='Input TSV file containing gene data')
    parser.add_argument('--rrna_file', help='rRNA TSV file')
    parser.add_argument('--trna_file', help='tRNA TSV file')
    parser.add_argument('--combined_annotations', help='Combined annotations TSV file')
    parser.add_argument('--combined_rrna', help='Combined rRNA TSV file')
    parser.add_argument('--output_file', help='Output XLSX file')
    args = parser.parse_args()

    generate_multi_sheet_xlsx(args.input_file, args.rrna_file, args.trna_file, args.combined_annotations, args.combined_rrna, args.output_file)